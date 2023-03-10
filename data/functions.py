import os
import json
import openpyxl
import calendar
import datetime

def update_additions(owner_id):
    # get estimator path from estimators.json
    estimators_file = os.path.join(os.getcwd(), 'data', 'estimators.json')
    with open(estimators_file, 'r') as f:
        estimators = json.load(f)
    estimator_path = estimators[owner_id]['estimator_path']

    # get price file path from accounts.json
    accounts_file = os.path.join(os.getcwd(), 'data', 'accounts.json')
    with open(accounts_file, 'r') as f:
        accounts = json.load(f)
    account = accounts[owner_id]
    price_file_path = account['price_file_path']

    # get current month name
    current_month_name = datetime.datetime.now().strftime("%B")

    # determine which columns to update based on current month
    if current_month_name == "January":
        update_columns = ['G', 'J']
    elif current_month_name == "February":
        update_columns = ['J', 'M']
    elif current_month_name == "March":
        update_columns = ['M', 'P']
    elif current_month_name == "April":
        update_columns = ['P', 'S']
    elif current_month_name == "May":
        update_columns = ['S', 'V']
    elif current_month_name == "June":
        update_columns = ['V', 'Y']
    elif current_month_name == "July":
        update_columns = ['Y', 'AB']
    elif current_month_name == "August":
        update_columns = ['AB', 'AE']
    elif current_month_name == "September":
        update_columns = ['AE', 'AH']
    elif current_month_name == "October":
        update_columns = ['AH', 'AK']
    elif current_month_name == "November":
        update_columns = ['AK', 'AN']
    else:
        update_columns = ['AN', 'G']

    # open price file additions excel file and get unique values in column A
    price_file_additions_path = os.path.join(estimator_path, 'Price File Additions.xlsx')
    wb_additions = openpyxl.load_workbook(price_file_additions_path)
    ws_additions = wb_additions.active
    account_numbers = set([cell.value for cell in ws_additions['A'][1:]])

    # open price file and update columns with values from price file additions
    wb_price_file = openpyxl.load_workbook(price_file_path)
    ws_price_file = wb_price_file.active
    last_row = ws_price_file.max_row

    for account_number in account_numbers:
        if account_number in accounts:
            account = accounts[account_number]
            price_file_path = account['price_file_path']
            wb_price_file = openpyxl.load_workbook(price_file_path)
            ws_price_file = wb_price_file.active
            for col in update_columns:
                for i in range(2, last_row + 1):
                    if ws_price_file['A'][i].value == account_number:
                        new_price = ws_additions['C'][ws_additions['A'].index(account_number)].value
                        if new_price is not None:
                            new_price = float(f'{new_price:.2f}')
                            ws_price_file[f'{col}{i}'].value = new_price

            wb_price_file.save(price_file_path)
            print(f'Price file for account {account_number} updated with values from price file additions.')

def update_price_files():
    pass

def create_new_price_file(owner_id):
    # Load existing data from accounts.json
    accounts_file = os.path.join(os.getcwd(), 'data', 'accounts.json')
    with open(accounts_file, 'r') as f:
        data = json.load(f)

    # Get nacet path from generalpaths.json
    generalpaths_file = os.path.join(os.getcwd(), 'data', 'generalpaths.json')
    with open(generalpaths_file, 'r') as f:
        general_paths = json.load(f)
    nacet_path = general_paths['nacet']

    # Get unique ID for the price file
    while True:
        unique_id = input("Enter unique ID of the price file: ")
        if unique_id in data:
            print(f"Price file with ID {unique_id} already exists. Please try again.")
            continue
        else:
            break

    # Get name of the price file
    while True:
        name = input("Enter name of the price file: ")
        name_exists = any(val['name'] == name for val in data.values())
        if name_exists:
            print(f"Price file with name {name} already exists. Please try again.")
            continue
        else:
            break

    # Get vertical for the price file
    print("Select the vertical for the price file:")
    print("1. New build")
    print("2. Social Housing")
    print("3. Private & Domestic")
    while True:
        vertical_input = input()
        if vertical_input not in ['1', '2', '3']:
            print("Invalid option. Please try again.")
            continue
        else:
            vertical = ['New build', 'Social Housing', 'Private & Domestic'][int(vertical_input)-1]
            break

    # Check if the price file has a copper file
    while True:
        copper_file_input = input("Does the price file have a copper file? (y/n): ")
        if copper_file_input.lower() == 'y':
            copper_file = True
            break
        elif copper_file_input.lower() == 'n':
            copper_file = False
            break
        else:
            print("Invalid option. Please try again.")
            continue

    # Get the price file type
    print("Select the price file type:")
    print("1. SLP")
    print("2. Margin sheet")
    while True:
        price_file_type_input = input()
        if price_file_type_input not in ['1', '2']:
            print("Invalid option. Please try again.")
            continue
        else:
            price_file_type = ['SLP', 'Margin sheet'][int(price_file_type_input)-1]
            break

    # Generate the price file path
    price_file_path = os.path.join(nacet_path, "Accounts", vertical, name, "Master Price File" , f"{unique_id} - {name} Master.xlsx")

    # Store the new data in accounts.json
    data[unique_id] = {
        "owner_id": owner_id,
        "name": name,
        "vertical": vertical,
        "copper_file": copper_file,
        "price_file_type": price_file_type,
        "price_file_path": price_file_path
    }
    with open(accounts_file, 'w') as f:
        json.dump(data, f, indent=2)

    print(f"Price file with ID {unique_id} is created.")

def list_all_accounts(user_id):
    filename = os.path.join(os.getcwd(), 'data', 'accounts.json')

    with open(filename, 'r') as f:
        data = json.load(f)

    user_accounts = []
    for unique_id, account in data.items():
        if account["owner_id"] == user_id:
            account["ID"] = unique_id
            user_accounts.append(account)

    if len(user_accounts) == 0:
        print("There are no accounts registered under your name.")
        return

    for account in user_accounts:
        print(f"{account['ID']}\t{account['name']}\t{account['vertical']}\t{account['price_file_type']}")

def create_new_user():
    estimators_file = os.path.join(os.getcwd(), 'data', 'estimators.json')
    with open(estimators_file, 'r') as f:
        data = json.load(f)

    generalpaths_file = os.path.join(os.getcwd(), 'data', 'generalpaths.json')
    with open(generalpaths_file, 'r') as f:
        general_paths = json.load(f)

    while True:
        first_name = input("Enter your first name (type 'exit' to cancel): ")
        if first_name.lower() == "exit":
            break
        elif first_name == "":
            print("First name cannot be blank. Please try again.")
            continue

        last_name = input("Enter your last name (type 'exit' to cancel): ")
        if last_name.lower() == "exit":
            break
        elif last_name == "":
            print("Last name cannot be blank. Please try again.")
            continue

        unique_id = (first_name[0] + last_name[0]).upper()

        if unique_id in data:
            print(f"Account with ID {unique_id} already exists. Please try again.")
            continue
        else:
            data[unique_id] = {"first_name": first_name, "last_name": last_name}

            # Use os.path.join() to construct the path using the appropriate directory separator
            estimator_path = os.path.join(general_paths["nacet"], "Estimators", f"{first_name} {last_name}")
            os.makedirs(estimator_path, exist_ok=True)
            data[unique_id]["estimator_path"] = estimator_path

            with open(estimators_file, 'w') as f:
                json.dump(data, f, indent=2)

            print(f"User {unique_id} is created.")
            break

def remove_user():
    filename = os.path.join(os.getcwd(), 'data', 'estimators.json')

    with open(filename, 'r') as f:
        data = json.load(f)

    while True:
        user_id = input("Enter user ID (type 'exit' to cancel): ")
        if user_id.lower() == "exit":
            break
        elif user_id not in data:
            print(f"User ID {user_id} not found. Please try again.")
            continue

        first_name = data[user_id]["first_name"]
        last_name = data[user_id]["last_name"]

        confirm = input(f"Are you sure you want to remove {user_id} {first_name} {last_name} from the database? (Yes/No) ")
        if confirm.lower() == "yes":
            del data[user_id]
            with open(filename, 'w') as f:
                json.dump(data, f, indent=2)
            print(f"User {user_id} is removed.")
            break
        elif confirm.lower() == "no":
            break
        else:
            print("Invalid choice. Please try again.")

def list_all_users():
    filename = os.path.join(os.getcwd(), 'data', 'estimators.json')

    with open(filename, 'r') as f:
        data = json.load(f)

    print("ID\t\tFirst Name\tLast Name")
    print("--------------------------------------------------")

    for user_id, user_data in data.items():
        first_name = user_data["first_name"]
        last_name = user_data["last_name"]
        print(f"{user_id}\t\t{first_name}\t\t{last_name}")


def export_data():
    accounts_file = os.path.join(os.getcwd(), 'data', 'accounts.json')
    estimators_file = os.path.join(os.getcwd(), 'data', 'estimators.json')
    wb_file = os.path.join(os.getcwd(), 'data', 'exported_data.xlsx')

    if os.path.exists(wb_file):
        os.remove(wb_file)

    wb = openpyxl.Workbook()

    default_sheet = wb.active
    default_sheet.title = "accounts"

    headers = ['Unique ID', 'Owner', 'Name', 'Vertical', 'Copperfile', 'Pricefiletype']
    default_sheet.append(headers)

    with open(accounts_file, 'r') as f:
        data = json.load(f)
        for unique_id, price_file in data.items():
            row = [unique_id, price_file['owner_id'], price_file['name'], price_file['vertical'], price_file['copper_file'], price_file['price_file_type']]
            default_sheet.append(row)

    wb.save(wb_file)

    print(f"Data exported to {wb_file}!")